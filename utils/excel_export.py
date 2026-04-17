import os
from datetime import datetime
from flask import current_app, send_file
import io
import logging

logger = logging.getLogger(__name__)

class ExcelExporter:
    """
    Export reports to Excel format.
    Uses openpyxl if available, otherwise falls back to CSV.
    """
    
    @staticmethod
    def export_sales_report(sales_data, filename=None):
        """
        Export sales list to Excel.
        sales_data: list of sale objects or dicts.
        """
        if not sales_data:
            return None
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Report"
            
            # Headers
            headers = ['Invoice', 'Date', 'Cashier', 'Patient', 'Subtotal', 'Tax', 'Discount', 'Total', 'Payment Method']
            ws.append(headers)
            
            # Style headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for sale in sales_data:
                row = [
                    sale.invoice_number,
                    sale.created_at.strftime('%Y-%m-%d %H:%M'),
                    sale.cashier.username,
                    sale.patient.full_name if sale.patient else 'Walk-in',
                    sale.subtotal,
                    sale.tax,
                    sale.discount,
                    sale.total,
                    sale.payment_method
                ]
                ws.append(row)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            if not filename:
                filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return output, filename
            
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return ExcelExporter._export_csv(sales_data, filename)
    
    @staticmethod
    def _export_csv(sales_data, filename=None):
        """Fallback CSV export."""
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow(['Invoice', 'Date', 'Cashier', 'Patient', 'Subtotal', 'Tax', 'Discount', 'Total', 'Payment Method'])
        
        for sale in sales_data:
            writer.writerow([
                sale.invoice_number,
                sale.created_at.strftime('%Y-%m-%d %H:%M'),
                sale.cashier.username,
                sale.patient.full_name if sale.patient else 'Walk-in',
                sale.subtotal,
                sale.tax,
                sale.discount,
                sale.total,
                sale.payment_method
            ])
        
        output.seek(0)
        if not filename:
            filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        # Convert to bytes for send_file
        bytes_output = io.BytesIO(output.getvalue().encode('utf-8'))
        return bytes_output, filename
    
    @staticmethod
    def export_stock_report(drugs, filename=None):
        """Export current stock to Excel."""
        if not drugs:
            return None
        
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stock Report"
            
            headers = ['Name', 'Generic Name', 'Category', 'Quantity', 'Reorder Level', 'Buying Price', 'Selling Price', 'Expiry Date', 'Supplier']
            ws.append(headers)
            
            for drug in drugs:
                ws.append([
                    drug.name,
                    drug.generic_name or '',
                    drug.category or '',
                    drug.quantity,
                    drug.reorder_level,
                    drug.buying_price,
                    drug.selling_price,
                    drug.expiry_date.strftime('%Y-%m-%d'),
                    drug.supplier.name if drug.supplier else ''
                ])
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            if not filename:
                filename = f"stock_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            return output, filename
        except ImportError:
            return None